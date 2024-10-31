# pip install pandas zipfile36
#D:\\IVASCOG\\Desktop\\python\\aplicacion_powerBi\\datos\\Reporte DepartamentoANTIOQUIA20241007_014406.csv
import pandas as pd
import openpyxl
import os as os

# df = openpyxl.load_workbook('D:\\IVASCOG\\Desktop\\python\\aplicacion_powerBi\\datos\\Reporte DepartamentoANTIOQUIA20241029_091459.xlsx')

# #print(df.sheetnames)

# # dict
# libro_por_hojas = {}
# # iteración de hojas
# for hoja in df.sheetnames:
#     libro_por_hojas[hoja] = pd.DataFrame(df[hoja].values)
#     #print(f'hoja: {hoja} lista')

# # sacar una tabla del dict
# edad_sexo = pd.DataFrame(libro_por_hojas['Edad y Sexo'])

# def definir_columnas(dataset):
#     # tomar la el primer registro para colocarlo como columna
#     columnas = list(dataset.iloc[0])
#     # Transponer el DataFrame y luego restablecer los índices
#     dataset.columns= columnas[0:] # Establecer la primera fila como nombres de columnas
#     # Eliminar la primera fila
#     dataset = dataset[1:]
#     # Restablecer el índice
#     dataset.reset_index(drop=True, inplace=True)


#     return dataset

    


# # print(definir_columnas(edad_sexo).head())

# data2 = pd.read_csv('D:\\IVASCOG\\Desktop\\python\\aplicacion_powerBi\\datos\\RAT NOMINAL.csv')

# print(data2)


 # leer todos los documentos de la carpeta y almacenarlos en un diccionario
def cargar_documentos(ruta):
    documentos_csv = {}
    documenos_xlsx = {}
    
    # Recorre todos los archivos en la carpeta
    for nombre_archivo in os.listdir(ruta):
        if nombre_archivo.endswith('.csv'):
            documentos_csv[nombre_archivo] = pd.read_csv(os.path.join(ruta, nombre_archivo))
        
        if nombre_archivo.endswith('xlsx') or nombre_archivo.endswith('xlx'):
            documenos_xlsx[nombre_archivo] = pd.read_excel(os.path.join(ruta, nombre_archivo))

            # iteración de hojas
            df = openpyxl.load_workbook(os.path.join(ruta, nombre_archivo))
            names = df.sheetnames
            for hoja in names:
                if not hoja.endswith('.xlsx'):
                    documenos_xlsx[hoja] = pd.DataFrame(df[hoja].values)

    return documentos_csv, documenos_xlsx

# eliminar archivos extras
def eliminar_archivos_extra(diccionario: dict):
    for key, value in diccionario.items():
        if key.endswith('.xlsx'):
            eliminado = {}
            eliminado[key] = diccionario[key]
    
    for llave, value in eliminado.items():
        del diccionario[llave]

    return diccionario

def main():
    # leer todos los documentos de la carpeta y almacenarlos en un diccionario
    #print('presiona la tecla 1 para limpiar carga rlos datos')
    ruta = 'D:\\IVASCOG\\Desktop\\python\\aplicacion_powerBi\\datos'
    
    # extraer los datos de la carpeta
    documentos_csv, documenos_xlsx = cargar_documentos(ruta)

    # eliminar archivo extra
    documenos_xlsx = eliminar_archivos_extra(documenos_xlsx)


    

    # print('docuemntos csv')
    # for clave, valor in documentos_csv.items():
    #     print(f'{clave}')

    # for key, value in documenos_xlsx.items():
    #     if key.endswith('.xlsx'):
    #         eliminado = {}
    #         eliminado[key] = documenos_xlsx[key]
    
    # for llave, value in eliminado.items():
    #     del documenos_xlsx[llave]
    

    # for llave, value in documenos_xlsx.items():
    #     print(llave)



    # for nombre_archivo in os.listdir(ruta):
    #     if nombre_archivo.endswith('.csv'):
    #         documentos_csv[nombre_archivo] = pd.read_csv(os.path.join(ruta, nombre_archivo))
        
    #     if nombre_archivo.endswith('xlsx') or nombre_archivo.endswith('xlx'):
    #         documenos_xlsx[nombre_archivo] = pd.read_excel(os.path.join(ruta, nombre_archivo))

    #         # iteración de hojas
    #         df = openpyxl.load_workbook(os.path.join(ruta, nombre_archivo))
    #         #print(df.sheetnames)
    #         for hoja in df.sheetnames:
    #             documenos_xlsx[hoja] = pd.DataFrame(df[hoja].values)
    
    # for clave, valor in documentos_csv.items():
    #     print(f'{clave}')

    # for key, value in documenos_xlsx.items():
    #     print(f'{value}')

        # if os.path.isfile(ruta_archivo):
        #     if ruta_archivo.endswith('.csv'):
        #         with open(ruta_archivo, 'r') as archivo:
        #             documentos[nombre_archivo] = pd.read_csv(archivo)
    # limpiar todos los documentos de acuerdo a su necesidad

    # convertir todos los documentos en .csv

    # subir los archivos a Power BI

    # crar la automatización semanal por medio de carpetas con nombres especificos


main()
# for clave, valor in libro_por_hojas.items():
#     print(f'{clave}: {valor}')

# data = pd.read_excel(df.active)
# print(data)